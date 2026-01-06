"""
Excel 索引文件处理工具
"""
from pathlib import Path


class ExcelIndexConverter:
    """Excel 索引文件转换器"""

    def __init__(self, excel_path):
        """
        初始化转换器

        Args:
            excel_path: Excel 文件路径
        """
        self.excel_path = Path(excel_path)
        self.data = []

    def load(self, sheet_name=None, skip_rows=0, entry_col=0, page_col=1):
        """
        加载 Excel 文件

        Args:
            sheet_name: 工作表名称（None 表示活动工作表）
            skip_rows: 跳过前 N 行（用于跳过标题行）
            entry_col: 词目所在列（0-based）
            page_col: 页码所在列（0-based）

        Returns:
            加载的数据行数
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("需要安装 openpyxl 库：pip3 install openpyxl")

        # 加载工作簿
        wb = openpyxl.load_workbook(str(self.excel_path), read_only=True)

        # 选择工作表
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # 读取数据
        self.data = []
        for i, row in enumerate(ws.iter_rows(min_row=skip_rows + 1, values_only=True)):
            # 跳过空行
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            # 确保有足够的列
            if len(row) <= max(entry_col, page_col):
                continue

            # 提取词目和页码
            entry = row[entry_col]
            page = row[page_col]

            # 跳过空词目
            if entry is None or str(entry).strip() == '':
                continue

            # 清理数据
            entry_str = str(entry).strip()
            page_str = str(page).strip() if page is not None else ''

            # 跳过无效行
            if not entry_str or not page_str:
                continue

            # 处理多页码情况（支持多种分隔符）
            # 支持的分隔符：顿号、逗号、分号、空格等
            import re
            # 匹配常见的分隔符：中文顿号、英文逗号、中文逗号、分号、空格
            pages = re.split(r'[、,，;；\s]+', page_str)

            # 为每个页码创建一条记录
            for single_page in pages:
                single_page = single_page.strip()
                # 验证页码是否有效（只包含数字）
                if single_page and re.match(r'^\d+$', single_page):
                    self.data.append((entry_str, single_page))
                elif single_page:
                    # 如果不是纯数字，保留原样（可能是特殊页码如"iii"）
                    self.data.append((entry_str, single_page))

        wb.close()
        return len(self.data)

    def save_as_txt(self, output_path, encoding='utf-8'):
        """
        保存为 TXT 文件

        Args:
            output_path: 输出文件路径
            encoding: 文件编码（默认 UTF-8）

        Returns:
            保存的行数
        """
        output_file = Path(output_path)

        with open(output_file, 'w', encoding=encoding) as f:
            for entry, page in self.data:
                # 使用制表符分隔
                f.write(f"{entry}\t{page}\n")

        return len(self.data)

    def get_preview(self, max_lines=20):
        """
        获取预览内容

        Args:
            max_lines: 最大预览行数

        Returns:
            预览字符串
        """
        preview_lines = []
        for entry, page in self.data[:max_lines]:
            preview_lines.append(f"{entry}\t{page}")

        preview = '\n'.join(preview_lines)
        if len(self.data) > max_lines:
            preview += f"\n\n... 共 {len(self.data)} 行"

        return preview

    def get_data_count(self):
        """获取数据行数"""
        return len(self.data)

    @staticmethod
    def get_sheet_names(excel_path):
        """
        获取 Excel 文件中的所有工作表名称

        Args:
            excel_path: Excel 文件路径

        Returns:
            工作表名称列表
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("需要安装 openpyxl 库：pip3 install openpyxl")

        wb = openpyxl.load_workbook(str(excel_path), read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names

    @staticmethod
    def detect_structure(excel_path, sheet_name=None, max_rows=10):
        """
        自动检测 Excel 文件结构

        Args:
            excel_path: Excel 文件路径
            sheet_name: 工作表名称
            max_rows: 检测的最大行数

        Returns:
            dict: {
                'has_header': bool,
                'entry_col': int,
                'page_col': int,
                'skip_rows': int,
                'sample_data': list
            }
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("需要安装 openpyxl 库：pip3 install openpyxl")

        wb = openpyxl.load_workbook(str(excel_path), read_only=True)

        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # 读取前几行数据
        sample_data = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True)):
            if row:
                sample_data.append(row)

        wb.close()

        if not sample_data:
            return None

        # 简单的启发式检测
        result = {
            'has_header': False,
            'entry_col': 0,
            'page_col': 1,
            'skip_rows': 0,
            'sample_data': sample_data
        }

        # 检测是否有标题行
        # 如果第一行包含"词目"、"页码"等关键词，认为有标题
        first_row = sample_data[0] if sample_data else []
        header_keywords = ['词目', '词条', '词', 'entry', 'word', '页码', '页', 'page']

        for i, cell in enumerate(first_row):
            if cell and any(keyword in str(cell).lower() for keyword in header_keywords):
                result['has_header'] = True
                result['skip_rows'] = 1
                break

        # 检测词目列和页码列
        # 通常第一列是词目，第二列是页码
        # 但也可能是其他顺序，这里简化处理
        if len(first_row) >= 2:
            result['entry_col'] = 0
            result['page_col'] = 1

        return result
